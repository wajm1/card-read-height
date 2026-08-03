#!/usr/bin/env python3
# ---------------------------------------------------------------------------
# Author:  Wajahat Mahmood
# Updated: 2026-07-30
# Project: rf IDEAS Credential Read Height Automation
# Summary: see the module docstring below for this file's responsibility.
# ---------------------------------------------------------------------------
"""format_results_xlsx.py — merge read-height CSVs into a formatted Excel report.

Matches the layout of results/Keep/2026-07-13_Read_Heights_Formatted.xlsx:
  title, metadata block, two-row angle headers, navy styling.

Usage (from Automation/):
  python tools/format_results_xlsx.py \\
      ../results/2026-07-23_09-14-45_HIP2_SP_read_heights.csv \\
      ../results/2026-07-23_09-47-42_HIP2_SP_read_heights.csv \\
      -o ../results/Keep/2026-07-23_HIP2_SP_Read_Heights_Formatted.xlsx

With no args, merges today's two HIP2_SP files in ../results/ if present.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
    _OPENPYXL_ERR = None
except ImportError as e:
    Workbook = None  # type: ignore
    _OPENPYXL_OK = False
    _OPENPYXL_ERR = e


def _require_openpyxl():
    if not _OPENPYXL_OK:
        raise ImportError(
            "openpyxl is required for formatted Excel reports. "
            "Install with: pip install openpyxl"
        ) from _OPENPYXL_ERR

# Mojibake / Excel-ANSI cleanup: strip UTF-8 misreads and normalize punctuation.
_MOJIBAKE_FIXES = (
    ("Â°", "°"),
    ("Â", ""),          # leftover C2 from mis-decoded UTF-8
    ("â€”", "—"),       # em dash misread as CP1252
    ("â€“", "–"),
    ("\u00c2\u00b0", "°"),
)


def _clean_text(value):
    if value is None:
        return ""
    s = str(value)
    for bad, good in _MOJIBAKE_FIXES:
        s = s.replace(bad, good)
    # Title: prefer ASCII hyphen in filenames; keep unicode dash in Excel cells
    return s.strip()


def _open_csv(path):
    last = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return open(path, newline="", encoding=enc)
        except UnicodeDecodeError as e:
            last = e
    raise last  # type: ignore[misc]


def _parse_csv(path):
    """Return (meta dict, list of data dicts keyed by header names)."""
    meta = {"sources": [os.path.basename(path)]}
    rows_out = []
    with _open_csv(path) as f:
        reader = csv.reader(f)
        header = None
        for raw in reader:
            if not raw:
                continue
            cells = [_clean_text(c) for c in raw]
            key = cells[0].lower()
            if key.startswith("rf ideas"):
                meta["title"] = cells[0]
                continue
            if key == "reader type":
                meta["reader_type"] = cells[1] if len(cells) > 1 else ""
                if len(cells) > 3 and cells[2].lower() == "reader model":
                    meta["reader_model"] = cells[3]
                continue
            if key == "firmware":
                meta["firmware"] = cells[1] if len(cells) > 1 else ""
                continue
            if key == "test speed":
                meta["test_speed"] = cells[1] if len(cells) > 1 else ""
                if len(cells) > 3:
                    meta["final_tap"] = cells[3]
                continue
            if key == "read angles":
                meta["read_angles"] = cells[1] if len(cells) > 1 else ""
                continue
            if key == "comment":
                meta["comment"] = cells[1] if len(cells) > 1 else ""
                continue
            if key == "generated":
                meta["generated"] = cells[1] if len(cells) > 1 else ""
                continue
            if key == "run" and len(cells) > 5:
                header = cells
                continue
            if header is None:
                continue
            # data row
            while len(cells) < len(header):
                cells.append("")
            rows_out.append({header[i]: cells[i] for i in range(len(header))})
    return meta, rows_out


def _num(value):
    s = _clean_text(value)
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return s


def _merge_metas(metas):
    out = {
        "reader_type": "",
        "reader_model": "",
        "firmware": "(not recorded)",
        "test_speed": "",
        "final_tap": "",
        "read_angles": "",
        "generated": datetime.now().strftime("%b-%d-%Y %H:%M:%S"),
        "sources": [],
    }
    angles = set()
    for m in metas:
        out["sources"].extend(m.get("sources") or [])
        for k in ("reader_type", "reader_model", "test_speed", "final_tap"):
            if m.get(k) and not out[k]:
                out[k] = m[k]
        fw = (m.get("firmware") or "").strip()
        if fw and out["firmware"] in ("", "(not recorded)"):
            out["firmware"] = fw
        for part in re.split(r"\s*,\s*", m.get("read_angles") or ""):
            if part:
                angles.add(part)
    if angles:
        # sort by numeric degree when possible
        def _ang_key(a):
            m = re.match(r"(\d+)", a)
            return int(m.group(1)) if m else 999

        out["read_angles"] = ", ".join(sorted(angles, key=_ang_key))
    return out


def build_workbook(metas, data_rows):
    _require_openpyxl()
    meta = _merge_metas(metas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Read Heights"

    navy = "1F3864"
    navy2 = "2F5597"
    fill_navy = PatternFill("solid", fgColor=navy)
    fill_sub = PatternFill("solid", fgColor=navy2)
    font_title = Font(name="Calibri", size=16, bold=True, color=navy)
    font_label = Font(name="Calibri", size=10, bold=True, color="404040")
    font_value = Font(name="Calibri", size=10, color="000000")
    font_hdr = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_sub = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="B0B0B0")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center")

    # Title
    ws["A1"] = "rf IDEAS Credential Read Height Test — Formatted Report"
    ws["A1"].font = font_title
    ws["A1"].alignment = left_mid
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 25.5

    # Metadata (label / value in A/B)
    meta_rows = [
        ("Reader Type", meta.get("reader_type") or ""),
        ("Reader Model", meta.get("reader_model") or meta.get("reader_type") or ""),
        ("Firmware", meta.get("firmware") or "(not recorded)"),
        ("Test Speed", meta.get("test_speed") or ""),
        ("Final Tap", meta.get("final_tap") or ""),
        ("Read Angles", meta.get("read_angles") or ""),
        ("Generated", meta.get("generated") or ""),
        ("Source files", "; ".join(meta.get("sources") or [])),
    ]
    for i, (lab, val) in enumerate(meta_rows, start=3):
        ws.cell(i, 1, lab).font = font_label
        ws.cell(i, 2, val).font = font_value
        ws.row_dimensions[i].height = 15

    # Header rows 11–12
    # A-C identity (no Side — barcode A###/B### is enough),
    # D-G 0°, H-K 90°, L-O 180°, P-S 270°, T Card Max, U Error
    headers_top = [
        (1, "Card #"), (2, "Card Title"), (3, "Card Code"),
        (4, "0° Reading"), (8, "90° Reading"), (12, "180° Reading"),
        (16, "270° Reading"), (20, "Card Max (mm)"), (21, "Error / Skip"),
    ]
    for col, text in headers_top:
        cell = ws.cell(11, col, text)
        cell.font = font_hdr
        cell.fill = fill_navy
        cell.alignment = center
        cell.border = border_all

    for col in range(1, 22):
        c11 = ws.cell(11, col)
        c12 = ws.cell(12, col)
        if c11.value is None:
            c11.fill = fill_navy
            c11.border = border_all
        c12.fill = fill_sub
        c12.font = font_sub
        c12.alignment = center
        c12.border = border_all

    for col, label in ((4, "Avg"), (5, "Min"), (6, "Max"), (7, "Scans"),
                       (8, "Avg"), (9, "Min"), (10, "Max"), (11, "Scans"),
                       (12, "Avg"), (13, "Min"), (14, "Max"), (15, "Scans"),
                       (16, "Avg"), (17, "Min"), (18, "Max"), (19, "Scans")):
        ws.cell(12, col, label)

    ws.merge_cells("A11:A12")
    ws.merge_cells("B11:B12")
    ws.merge_cells("C11:C12")
    ws.merge_cells("D11:G11")
    ws.merge_cells("H11:K11")
    ws.merge_cells("L11:O11")
    ws.merge_cells("P11:S11")
    ws.merge_cells("T11:T12")
    ws.merge_cells("U11:U12")
    ws.row_dimensions[11].height = 19.5
    ws.row_dimensions[12].height = 15.75

    # Column map from CSV header names → Excel columns
    # CSV: Run, Card #, Card Title, Card Code,  (Side omitted)
    #   0° Avg, 90° Avg, 180° Avg, 270° Avg,
    #   0° Min, 0° Max, 90° Min, 90° Max, 180° Min, 180° Max, 270° Min, 270° Max,
    #   0° Scans, 90° Scans, 180° Scans, 270° Scans, Card Max, Error
    def _get(row, *names):
        for n in names:
            for k, v in row.items():
                if _clean_text(k).lower() == n.lower():
                    return v
        return ""

    start = 13
    for i, row in enumerate(data_rows):
        r = start + i
        values = [
            _num(_get(row, "Card #")) or _get(row, "Card #"),
            _get(row, "Card Title"),
            _get(row, "Card Code"),
            _num(_get(row, "0° Avg (mm)", "0 deg Avg (mm)")),
            _num(_get(row, "0° Min", "0 deg Min")),
            _num(_get(row, "0° Max", "0 deg Max")),
            _num(_get(row, "0° Scans", "0 deg Scans")),
            _num(_get(row, "90° Avg (mm)", "90 deg Avg (mm)")),
            _num(_get(row, "90° Min", "90 deg Min")),
            _num(_get(row, "90° Max", "90 deg Max")),
            _num(_get(row, "90° Scans", "90 deg Scans")),
            _num(_get(row, "180° Avg (mm)", "180 deg Avg (mm)")),
            _num(_get(row, "180° Min", "180 deg Min")),
            _num(_get(row, "180° Max", "180 deg Max")),
            _num(_get(row, "180° Scans", "180 deg Scans")),
            _num(_get(row, "270° Avg (mm)", "270 deg Avg (mm)")),
            _num(_get(row, "270° Min", "270 deg Min")),
            _num(_get(row, "270° Max", "270 deg Max")),
            _num(_get(row, "270° Scans", "270 deg Scans")),
            _num(_get(row, "Card Max (mm)")),
            _get(row, "Error / Skip"),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(r, c, "" if val is None else val)
            cell.font = font_data
            cell.border = border_all
            if c == 1 or c >= 4:
                cell.alignment = center
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 15

    widths = {
        1: 8, 2: 24, 3: 12,
        4: 7, 5: 7, 6: 7, 7: 7,
        8: 7, 9: 7, 10: 7, 11: 7,
        12: 7, 13: 7, 14: 7, 15: 7,
        16: 7, 17: 7, 18: 7, 19: 7,
        20: 12, 21: 16,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def format_read_heights_csv(csv_path, xlsx_path=None):
    """Build the formatted Excel report from one read-heights CSV.

    Returns the output .xlsx path. Raises ImportError if openpyxl is missing,
    OSError/ValueError on bad input.
    """
    if not os.path.isfile(csv_path):
        raise FileNotFoundError(csv_path)
    meta, rows = _parse_csv(csv_path)
    if xlsx_path is None:
        base, _ = os.path.splitext(csv_path)
        xlsx_path = base + "_Formatted.xlsx"
    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)) or ".", exist_ok=True)
    wb = build_workbook([meta], rows)
    wb.save(xlsx_path)
    return xlsx_path


def main(argv=None):
    if not _OPENPYXL_OK:
        print("Install openpyxl:  pip install openpyxl", file=sys.stderr)
        return 1
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("csvs", nargs="*", help="Input read_heights CSV paths")
    ap.add_argument("-o", "--output", help="Output .xlsx path")
    args = ap.parse_args(argv)

    root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    results = os.path.join(root, "results")
    csvs = list(args.csvs)
    if not csvs:
        defaults = [
            os.path.join(results, "2026-07-23_09-14-45_HIP2_SP_read_heights.csv"),
            os.path.join(results, "2026-07-23_09-47-42_HIP2_SP_read_heights.csv"),
        ]
        csvs = [p for p in defaults if os.path.isfile(p)]
        if not csvs:
            print("Pass one or more CSV paths, or place files in results/.",
                  file=sys.stderr)
            return 2
        if len(csvs) == 1:
            out = args.output or (os.path.splitext(csvs[0])[0] + "_Formatted.xlsx")
            path = format_read_heights_csv(csvs[0], out)
            print("Wrote {} ({} data rows)".format(path, "see file"))
            return 0

    metas, all_rows = [], []
    for path in csvs:
        if not os.path.isfile(path):
            print("Missing:", path, file=sys.stderr)
            return 2
        meta, rows = _parse_csv(path)
        metas.append(meta)
        all_rows.extend(rows)

    out = args.output or os.path.join(
        results, "Keep",
        "{}_HIP2_SP_Read_Heights_Formatted.xlsx".format(
            datetime.now().strftime("%Y-%m-%d"),
        ),
    )
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb = build_workbook(metas, all_rows)
    wb.save(out)
    print("Wrote {} ({} data rows from {} file(s))".format(
        out, len(all_rows), len(csvs)))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
