# ---------------------------------------------------------------------------
# Author:  Wajahat Mahmood
# Updated: 2026-07-30
# Project: rf IDEAS Credential Read Height Automation
# Summary: see the module docstring below for this file's responsibility.
# ---------------------------------------------------------------------------
﻿import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

base = Path(r"C:\Users\wmahmood\OneDrive - rfIDEAS\Documents\card-read-heights\card-read-height")
xlsx = base / "results" / "Keep" / "2026-07-13_Read_Heights_Formatted.xlsx"
out_path = base / "_inspect_output.txt"

lines = []
def p(*a):
    s = " ".join(str(x) for x in a)
    lines.append(s)
    print(s)

wb = load_workbook(xlsx)
ws = wb.active
p("=== WORKBOOK ===")
p("Active sheet:", ws.title)
p("Sheet names:", wb.sheetnames)
p("Dimensions:", ws.dimensions)
p()

p("=== CELL VALUES rows 1-20, cols A-V ===")
for r in range(1, 21):
    vals = []
    for c in range(1, 23):
        v = ws.cell(r, c).value
        if v is None:
            vals.append("")
        else:
            vals.append(repr(v) if not isinstance(v, str) else v)
    p("R%02d: %s" % (r, " | ".join(vals)))
p()

p("=== MERGED CELL RANGES ===")
for m in ws.merged_cells.ranges:
    p(str(m))
p()

def describe_cell(addr):
    cell = ws[addr]
    font = cell.font
    fill = cell.fill
    align = cell.alignment
    p("--- %s ---" % addr)
    p("  value: %r" % (cell.value,))
    p("  font: name=%r size=%s bold=%s italic=%s color=%s" % (font.name, font.size, font.bold, font.italic, font.color))
    if font.color is not None:
        p("    font.color.type=%s rgb=%s theme=%s indexed=%s" % (font.color.type, getattr(font.color, "rgb", None), getattr(font.color, "theme", None), getattr(font.color, "indexed", None)))
    p("  fill: patternType=%s fgColor=%s bgColor=%s" % (fill.patternType, fill.fgColor, fill.bgColor))
    if fill.fgColor is not None:
        fc = fill.fgColor
        p("    fgColor.type=%s rgb=%s theme=%s indexed=%s tint=%s" % (fc.type, getattr(fc, "rgb", None), getattr(fc, "theme", None), getattr(fc, "indexed", None), getattr(fc, "tint", None)))
    p("  alignment: horizontal=%r vertical=%r wrap_text=%s shrink_to_fit=%s indent=%s" % (align.horizontal, align.vertical, align.wrap_text, align.shrink_to_fit, align.indent))
    p("  number_format: %r" % (cell.number_format,))
    p("  border: left=%s right=%s top=%s bottom=%s" % (cell.border.left.style, cell.border.right.style, cell.border.top.style, cell.border.bottom.style))

p("=== FONT/FILL/ALIGNMENT KEY CELLS ===")
for addr in ["A1", "A3", "A11", "E11", "E12"]:
    describe_cell(addr)
p()

p("=== COLUMN WIDTHS A-V ===")
for c in range(1, 23):
    letter = get_column_letter(c)
    w = ws.column_dimensions[letter].width
    p("  %s: %s" % (letter, w))
p()
p("=== ROW HEIGHTS 1-20 ===")
for r in range(1, 21):
    h = ws.row_dimensions[r].height
    p("  row %s: %s" % (r, h))
p()

def parse_csv(path):
    p("=== CSV: %s ===" % path.name)
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    p("Total rows in file: %s" % len(rows))
    p("--- ALL RAW ROWS ---")
    for i, row in enumerate(rows, 1):
        p("  L%03d: %s" % (i, row))
    run_idx = None
    for i, row in enumerate(rows):
        if row and row[0].strip().lower() == "run":
            run_idx = i
            break
    if run_idx is None:
        for i, row in enumerate(rows):
            if row and any(c.strip().lower() == "run" for c in row if c):
                run_idx = i
                break
    one_based = None if run_idx is None else run_idx + 1
    p("--- Run header at line index %s (1-based %s) ---" % (run_idx, one_based))
    if run_idx is not None:
        p("METADATA (before Run header):")
        for row in rows[:run_idx]:
            p("  %s" % (row,))
        p("HEADER: %s" % (rows[run_idx],))
        data = rows[run_idx + 1 :]
        p("DATA ROWS (%s):" % len(data))
        for i, row in enumerate(data, 1):
            p("  D%03d: %s" % (i, row))
    p()

for name in [
    "2026-07-23_09-14-45_HIP2_SP_read_heights.csv",
    "2026-07-23_09-47-42_HIP2_SP_read_heights.csv",
]:
    parse_csv(base / "results" / name)

out_path.write_text("\n".join(lines), encoding="utf-8")
p("WROTE", out_path)
